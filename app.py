import streamlit as st
import google.generativeai as genai
from bs4 import BeautifulSoup
import requests
import pandas as pd
import json
import time
import io

# --- Configuración de la Página ---
st.set_page_config(
    page_title="AI Web Auditor · UHU",
    page_icon="🕵️‍♂️",
    layout="wide"
)

# --- Estilos Corporativos UHU ---
def set_uhu_styles():
    import streamlit as st
    st.markdown("""
    <link href="https://fonts.googleapis.com/css2?family=Open+Sans:wght@300;400;600;700&display=swap" rel="stylesheet">
    <style>
    /* Colores UHU: Burdeos (#a90a2e), Negro (#000000), Blanco (#ffffff) */
    :root {
        --uhu-burdeos: #a90a2e;
        --uhu-blanco: #ffffff;
        --uhu-negro: #000000;
        --uhu-gris: #f4f4f4;
    }
    html, body, [class*="css"] {
        font-family: 'Open Sans', 'Helvetica Neue', Helvetica, Arial, sans-serif !important;
    }
    button[kind="primary"] {
        background-color: var(--uhu-burdeos) !important;
        color: var(--uhu-blanco) !important;
        border: None !important;
        font-weight: bold !important;
        transition: all 0.3s ease !important;
    }
    button[kind="primary"]:hover {
        background-color: var(--uhu-negro) !important;
        border: None !important;
        transform: scale(1.02);
    }
    h1, h2, h3, h4 {
        color: var(--uhu-burdeos) !important;
    }
    [data-testid="stMetricValue"] {
        color: var(--uhu-burdeos) !important;
    }
    [data-testid="stSidebarUserContent"] h1,
    [data-testid="stSidebarUserContent"] h2,
    [data-testid="stSidebarUserContent"] h3 {
        color: var(--uhu-burdeos) !important;
    }
    .stProgress > div > div > div > div {
        background-color: var(--uhu-burdeos) !important;
    }
    div[data-testid="stSidebar"] {
        border-right: 3px solid var(--uhu-burdeos);
    }
    </style>
    """, unsafe_allow_html=True)

# --- Checklists ACCUA predefinidas ---
CHECKLISTS_ACCUA = {
    "Grado / Máster — Dimensión 1: Información Pública": [
        "1.1.a La web publica la denominación completa del título, menciones (Grado) o especialidades (Máster), universidades en caso de título conjunto, modalidad, número total de créditos, idioma/s de impartición y número de plazas ofertadas",
        "1.1.b La web publica los requisitos de acceso y procedimientos de admisión de estudiantes",
        "1.1.c La web publica los criterios de reconocimiento y transferencia de créditos",
        "1.1.d La web publica información sobre los programas de movilidad de los estudiantes propios y de acogida",
        "1.1.e La web publica los perfiles de ingreso y de egreso a los que se orientan las enseñanzas",
        "1.1.f La web publica el plan de estudios completo: módulos/materias/asignaturas, número de créditos ECTS, tipología (básica, obligatoria, optativa, PAE), información sobre TFG/TFM y organización temporal",
        "1.1.g La web publica la descripción de actividades y metodologías docentes y los sistemas de evaluación para cada asignatura, incluyendo contenidos docentes e información sobre el profesorado y la persona coordinadora",
        "1.1.h La web publica el perfil básico del profesorado: número de profesores/as, número de doctores/as, categorías y acreditaciones, quinquenios, sexenios y áreas de conocimiento",
        "1.1.i La web publica los medios materiales y servicios disponibles: espacios docentes, instalaciones, laboratorios, aulas informáticas",
        "1.1.j La web publica información sobre prácticas académicas externas: mecanismo de organización, criterios de elección, convenios y número de plazas disponibles (si aplica)",
        "1.1.k La web publica información sobre las acciones de apoyo y orientación académica y profesional del estudiantado",
        "1.1.l La información aportada en la web del título es clara y fácilmente accesible para el estudiantado y la sociedad en su conjunto",
        "1.1.m La información del título es accesible a personas con diversidad funcional",
        "1.1.n Toda la información del título está contenida en una única página web, o si existen varias, se garantiza la homogeneidad y actualización simultánea de todas ellas",
        "1.2.a La web publica los resultados de satisfacción de todos los grupos de interés: estudiantes, PDI, personal de apoyo, egresados y empleadores",
        "1.2.b La web publica datos y resultados del título: oferta y demanda académica, resultados por asignaturas y globales, datos del estudiantado, personal académico y empleabilidad",
        "1.3.a La web da acceso al Sistema de Garantía Interna de Calidad (SGIC): responsables, procedimientos y acciones de mejora puestas en marcha",
        "1.3.b La web da acceso a la Memoria de Verificación del título (y la versión modificada si aplica)",
        "1.3.c La web da acceso a los informes de seguimiento y de renovación de la acreditación del título",
        "1.3.d La web da acceso al Plan de Mejora del título",
        "1.3.e La web incluye las normativas académicas aplicables: matrícula, acceso y admisión, permanencia, reconocimiento y transferencia de créditos, movilidad, evaluación, prácticas externas y TFG/TFM"
    ],
    "Grado / Máster — Dimensión 2: El Título": [
        "2.1 Los objetivos del título se ajustan a las necesidades actuales del alumnado, del mercado laboral y a las demandas de la sociedad",
        "2.2 El perfil de egreso del título es consistente con el nivel del Marco Español de Cualificaciones para la Educación Superior (MECES)",
        "2.3 El plan de estudios está diseñado de acuerdo con el perfil de egreso y permite alcanzar los resultados de aprendizaje previstos",
        "2.4 El plan de estudios se desarrolla en la práctica de acuerdo con lo establecido en la memoria verificada",
        "2.5 Los criterios de acceso y admisión son coherentes con el perfil de ingreso y se aplican de forma transparente"
    ],
    "Grado / Máster — Dimensión 3: Recursos Humanos": [
        "3.1 El profesorado que imparte docencia en el título es suficiente en número y adecuado en cuanto a cualificación y dedicación para las materias que imparte",
        "3.2 El profesorado es evaluado y recibe apoyo para el desarrollo de su labor docente e investigadora",
        "3.3 El personal de apoyo es suficiente en número y tiene la cualificación adecuada para la atención de las necesidades del título"
    ],
    "Grado / Máster — Dimensión 4: Recursos Materiales": [
        "4.1 Los recursos materiales y servicios disponibles son adecuados para el desarrollo de las actividades formativas planificadas para el título",
        "4.2 Los convenios de prácticas externas son suficientes y adecuados para la consecución de las competencias previstas del título (si aplica)"
    ],
    "Grado / Máster — Dimensión 5: Resultados del Aprendizaje": [
        "5.1 Los resultados de aprendizaje previstos son alcanzados por el estudiantado que supera las asignaturas",
        "5.2 Los TFG/TFM evidencian la adquisición de los resultados de aprendizaje del título"
    ],
    "Grado / Máster — Dimensión 6: Calidad de la Docencia": [
        "6.1 La valoración del profesorado por parte del estudiantado, reflejada en las encuestas de satisfacción, es positiva",
        "6.2 Los sistemas de apoyo al aprendizaje son adecuados y el alumnado hace un uso positivo de ellos"
    ],
    "Grado / Máster — Dimensión 7: Indicadores y Mejora": [
        "7.1 Las tasas de rendimiento, éxito y eficiencia del título son adecuadas",
        "7.2 La tasa de graduación del título es adecuada o muestra una tendencia de mejora",
        "7.3 La tasa de abandono del título es adecuada o muestra una tendencia de mejora",
        "7.4 Los indicadores son analizados y se han puesto en marcha acciones de mejora pertinentes"
    ]
}

# --- Funciones Auxiliares ---

from urllib.parse import urljoin, urlparse

@st.cache_data(show_spinner=False, ttl=3600)
def scrape_single_page(url):
    """Auxiliar para scrapear una sola página y obtener texto + links. Cacheada 1h."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        for script in soup(["script", "style", "footer", "nav", "header", "noscript", "meta"]):
            script.decompose()

        internal_links = []
        base_domain = urlparse(url).netloc
        
        for a in soup.find_all('a', href=True):
            link_text = a.get_text(strip=True)
            link_href = a['href']
            full_url = urljoin(url, link_href)
            
            if urlparse(full_url).netloc == base_domain:
                internal_links.append(full_url)

            replacement = f" [ENLACE: {link_text} | URL: {link_href}] "
            a.replace_with(replacement)

        text = soup.get_text(separator=' ')
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        clean_text = '\n'.join(chunk for chunk in chunks if chunk)
        
        return clean_text, internal_links
        
    except Exception as e:
        return f"Error en {url}: {e}", []

def crawl_website(start_url, max_pages=10):
    """
    Navega por el sitio web comenzando en start_url hasta max_pages.
    """
    visited = set()
    queue = [start_url]
    all_content = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    pages_scraped = 0
    
    while queue and pages_scraped < max_pages:
        url = queue.pop(0)
        if url in visited:
            continue
            
        visited.add(url)
        status_text.text(f"Analizando página {pages_scraped + 1}/{max_pages}: {url}")
        
        content, links = scrape_single_page(url)
        
        if content and not content.startswith("Error"):
            all_content.append(f"\n\n--- INICIO PÁGINA: {url} ---\n{content}\n--- FIN PÁGINA: {url} ---\n")
            for link in links:
                if link not in visited and link not in queue:
                    queue.append(link)
            pages_scraped += 1
            progress_bar.progress(pages_scraped / max_pages)
        
        time.sleep(0.2)  # Pausa reducida: el timeout de 10s ya garantiza respeto al servidor
        
    status_text.empty()
    progress_bar.empty()
    
    return "\n".join(all_content)

@st.cache_data(show_spinner=False, ttl=600)
def get_available_models(api_key):
    """Obtiene los modelos disponibles. Cacheado 10 min para evitar llamadas repetidas."""
    try:
        genai.configure(api_key=api_key)
        models = []
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                models.append(m.name)
        return models
    except Exception as e:
        return []

def analyze_compliance(text, checklist_items, api_key, model_name):
    """
    Analiza el cumplimiento de los requisitos utilizando el modelo seleccionado.
    """
    try:
        genai.configure(api_key=api_key)
        
        model = genai.GenerativeModel(model_name)
        
        checklist_str = "\n".join([f"- {item}" for item in checklist_items])
        
        prompt = f"""
        Actúa como un Auditor de Calidad Web experto en el marco de los procesos de renovación de la acreditación de títulos universitarios (ACCUA/REACU/ANECA). Tu tarea es verificar si el contenido de un sitio web cumple con una lista de requisitos de información pública.
        El contenido proporcionado puede incluir múltiples páginas del mismo sitio, separadas por marcadores de "INICIO PÁGINA" y "FIN PÁGINA".

        ### CONTENIDO DEL SITIO WEB (Rastreo de múltiples páginas):
        {text[:500000]} 
        (Texto truncado a 500k caracteres. Si el sitio es enorme, prioriza la información de las primeras páginas).

        ### LISTA DE REQUISITOS (CHECKLIST):
        {checklist_str}

        ### INSTRUCCIONES:
        1. Analiza CADA requisito contra TODO el contenido proporcionado.
        2. Si encuentras evidencia en CUALQUIER página, el requisito se considera CUMPLE.
        3. Si no hay evidencia clara pero hay indicios parciales, usa PARCIAL.
        4. Presta especial atención a los enlaces a archivos (marcados como [ENLACE: ...]) para requisitos que pidan documentos.
        5. Cada criterio tiene un código como (1.1.a, 1.2.b, etc.). Inclúyelo en tu respuesta si aparece.
        
        Debes responder EXCLUSIVAMENTE con un JSON válido. No incluyas bloques de código markdown, solo el JSON crudo.
        
        El formato del JSON debe ser una lista de objetos:
        [
            {{
                "item": "Texto del requisito original (incluyendo su código si lo tiene)",
                "status": "CUMPLE" | "NO CUMPLE" | "PARCIAL" | "NO APLICA",
                "evidencia": "Cita el texto exacto o describe el archivo/enlace encontrado. Si es posible, menciona en qué URL se encontró.",
                "confianza": "alta" | "media" | "baja"
            }},
            ...
        ]
        """
        
        response = model.generate_content(prompt)
        
        response_text = response.text.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
            
        return json.loads(response_text)
        
    except Exception as e:
         return {"error": str(e)}

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def generate_styled_excel(df, titulo="", url=""):
    """Genera un archivo Excel con formato corporativo UHU."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría ACCUA"

    # --- Encabezado portada ---
    header_font_white = Font(bold=True, color="FFFFFF", size=13)
    burdeos_fill = PatternFill(start_color="a90a2e", end_color="a90a2e", fill_type="solid")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Universidad de Huelva — Auditoría Web de Calidad ACCUA"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].fill = burdeos_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Título analizado: {titulo}" if titulo else "Título: No especificado"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].fill = PatternFill(start_color="f0d0d5", end_color="f0d0d5", fill_type="solid")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells('A3:D3')
    ws['A3'] = f"URL analizada: {url}" if url else ""
    ws['A3'].font = Font(italic=True, size=10)
    ws['A3'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    ws.append([])  # fila en blanco

    # --- Cabecera tabla ---
    headers = ["Criterio / Requisito", "Estado", "Evidencia / Justificación", "Confianza"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = header_font_white
        cell.fill = burdeos_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 25

    # Colores de estado corporativos UHU
    green_fill = PatternFill(start_color="e6c7ce", end_color="e6c7ce", fill_type="solid")
    green_font = Font(color="a90a2e", bold=True)
    
    red_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
    red_font = Font(color="8B0000", bold=True)
    
    yellow_fill = PatternFill(start_color="ebebeb", end_color="ebebeb", fill_type="solid")
    yellow_font = Font(color="444444")

    for _, row in df.iterrows():
        status_val = row.get('status', 'INCIERTO').upper()
        
        if 'CUMPLE' in status_val and 'NO' not in status_val:
            display_status = "✅ CUMPLE"
            fill, fnt = green_fill, green_font
        elif 'NO CUMPLE' in status_val:
            display_status = "❌ NO CUMPLE"
            fill, fnt = red_fill, red_font
        elif 'PARCIAL' in status_val:
            display_status = "⚠️ PARCIAL"
            fill, fnt = yellow_fill, yellow_font
        else:
            display_status = status_val
            fill, fnt = None, None

        ws.append([
            row.get('item', ''),
            display_status,
            row.get('evidencia', ''),
            row.get('confianza', '')
        ])
        
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 40
        if fill:
            cell_status = ws.cell(row=current_row, column=2)
            cell_status.fill = fill
            cell_status.font = fnt
            cell_status.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in [1, 3]:
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf_report(df, titulo="", url="", compliance_rate=0):
    """Genera un PDF con portada corporativa UHU y tabla de resultados."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        # Colores UHU
        UHU_BURDEOS = colors.HexColor("#a90a2e")
        UHU_BURDEOS_LIGHT = colors.HexColor("#f0d0d5")
        UHU_NEGRO = colors.HexColor("#000000")
        WHITE = colors.white
        GRAY = colors.HexColor("#ebebeb")

        styles = getSampleStyleSheet()
        style_center = ParagraphStyle('center', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica')
        style_left = ParagraphStyle('left', parent=styles['Normal'], alignment=TA_LEFT, fontName='Helvetica', fontSize=8)
        
        elems = []

        # Portada
        elems.append(Spacer(1, 1.5*cm))
        elems.append(Table([[Paragraph('<font color="white" size="18"><b>Informe de Auditoría Web de Calidad</b></font>', style_center)]],
                           colWidths=[16.6*cm],
                           style=TableStyle([
                               ('BACKGROUND', (0,0), (-1,-1), UHU_BURDEOS),
                               ('TOPPADDING', (0,0), (-1,-1), 14),
                               ('BOTTOMPADDING', (0,0), (-1,-1), 14),
                               ('ROUNDEDCORNERS', [8]),
                           ])))
        elems.append(Spacer(1, 0.5*cm))
        elems.append(Table([[Paragraph('<font color="white" size="13">Universidad de Huelva — ACCUA</font>', style_center)]],
                           colWidths=[16.6*cm],
                           style=TableStyle([('BACKGROUND', (0,0), (-1,-1), UHU_NEGRO)])))
        elems.append(Spacer(1, 1.0*cm))

        # Datos del título
        data_portada = [
            [Paragraph(f'<b>Título analizado:</b> {titulo or "No especificado"}', style_left)],
            [Paragraph(f'<b>URL analizada:</b> {url or "No especificada"}', style_left)],
            [Paragraph(f'<b>Fecha del informe:</b> {time.strftime("%d/%m/%Y %H:%M")}', style_left)],
            [Paragraph(f'<b>Tasa de cumplimiento:</b> {compliance_rate:.1f}%', style_left)],
        ]
        elems.append(Table(data_portada, colWidths=[16.6*cm],
                           style=TableStyle([
                               ('BACKGROUND', (0,0), (-1,-1), UHU_BURDEOS_LIGHT),
                               ('TOPPADDING', (0,0), (-1,-1), 6),
                               ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                               ('LEFTPADDING', (0,0), (-1,-1), 12),
                               ('BOX', (0,0), (-1,-1), 0.5, UHU_BURDEOS),
                           ])))
        elems.append(Spacer(1, 1.0*cm))

        # Tabla de resultados
        elems.append(Paragraph('<b>Resultados de la Auditoría</b>', ParagraphStyle('h2', parent=styles['Heading2'], textColor=UHU_BURDEOS)))
        elems.append(Spacer(1, 0.4*cm))

        table_data = [['Criterio / Requisito', 'Estado', 'Evidencia', 'Confianza']]
        for _, row in df.iterrows():
            status = row.get('status', '').upper()
            if 'CUMPLE' in status and 'NO' not in status:
                display = '✅ CUMPLE'
            elif 'NO CUMPLE' in status:
                display = '❌ NO CUMPLE'
            elif 'PARCIAL' in status:
                display = '⚠️ PARCIAL'
            else:
                display = status

            table_data.append([
                Paragraph(str(row.get('item', ''))[:120], style_left),
                Paragraph(f'<b>{display}</b>', ParagraphStyle('c', parent=style_left, alignment=TA_CENTER)),
                Paragraph(str(row.get('evidencia', ''))[:300], style_left),
                Paragraph(str(row.get('confianza', '')), style_left)
            ])

        col_widths = [7*cm, 2.5*cm, 5.5*cm, 1.6*cm]
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), UHU_BURDEOS),
            ('TEXTCOLOR', (0,0), (-1,0), WHITE),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor("#cccccc")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, GRAY]),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
        ])
        elems.append(Table(table_data, colWidths=col_widths, style=table_style, repeatRows=1))

        doc.build(elems)
        buffer.seek(0)
        return buffer
    except Exception as e:
        return None


# =====================================================
# --- Interfaz de Usuario ---
# =====================================================

set_uhu_styles()

# --- Sidebar ---
with st.sidebar:
    st.markdown("""
    <div style="background:#a90a2e;padding:18px 10px;border-radius:6px;text-align:center;margin-bottom:10px;">
        <div style="color:white;font-size:1.1em;font-weight:700;letter-spacing:1px;">🎓 Universidad de Huelva</div>
        <div style="color:#f8d0d7;font-size:0.78em;margin-top:4px;">Auditoría Web · ACCUA</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.header("⚙️ Configuración")
    api_key = st.text_input("Gemini API Key", type="password", help="Introduce tu clave API de Google Gemini.")
    
    selected_model = "models/gemini-1.5-flash"
    if api_key:
        available_models = get_available_models(api_key)
        if available_models:
            default_index = 0
            for i, m in enumerate(available_models):
                if 'flash' in m:
                    default_index = i
                    break
            selected_model = st.selectbox("Seleccionar Modelo", available_models, index=default_index)
        else:
            st.error("No se pudieron cargar los modelos. Verifica tu API Key.")
            st.warning("Usando modelo por defecto.")

    st.divider()
    st.markdown("### 🕷️ Rastreo Web")
    max_pages = st.slider("Máx. páginas a analizar", min_value=1, max_value=20, value=5,
                          help="Número máximo de páginas internas que el bot visitará.")

    st.markdown("---")
    st.markdown("### ℹ️ Sobre la app")
    st.info("""
    **AI Web Auditor** analiza la web de tu titulación contra los criterios de calidad de ACCUA usando IA (Google Gemini).
    
    Útil para la **Renovación de la Acreditación** de títulos Grado, Máster y Doctorado en Andalucía.
    """)

# --- Área Principal ---
st.title("🕵️‍♂️ AI Web Auditor · ACCUA")
st.markdown("Verifica el cumplimiento de los requisitos de calidad ACCUA en la web de tu titulación en segundos.")

# --- Disclaimer ---
st.markdown("""
<div style="
    background: #fff8e6;
    border-left: 5px solid #a90a2e;
    border-radius: 4px;
    padding: 12px 16px;
    margin: 12px 0 20px 0;
    font-size: 0.85em;
    color: #333;
    line-height: 1.5;
">
<strong>⚠️ Aviso Legal y Limitaciones de uso</strong><br>
Esta herramienta utiliza Inteligencia Artificial (Google Gemini) para analizar automáticamente el contenido público de sitios web universitarios. 
Los resultados generados son <strong>meramente orientativos</strong> y no tienen carácter vinculante ni oficial. 
No sustituyen la valoración experta de los evaluadores de la Agencia Andaluza del Conocimiento (ACCUA/DEVA) ni de ninguna otra agencia de calidad.<br>
<span style="color:#666;">La Universidad de Huelva no se responsabiliza de las decisiones tomadas basándose exclusivamente en los resultados de esta herramienta. 
El análisis está condicionado por el contenido accesible públicamente y por las páginas rastreadas, pudiendo existir información no detectada.</span>
</div>
""", unsafe_allow_html=True)

# --- Inputs principales ---
col1, col2 = st.columns([2, 1])
with col1:
    url_input = st.text_input("🌐 URL del Sitio Web del Título", placeholder="https://www.uhu.es/titulacion/...")
with col2:
    titulo_nombre = st.text_input("📋 Nombre del Título (opcional)", placeholder="Ej: Grado en Enfermería")

# --- Selector de Checklist ---
st.markdown("### 📝 Checklist de Requisitos")
modo_checklist = st.radio("Fuente de la checklist:", 
                           ["📋 Checklist ACCUA predefinida", "📁 Cargar fichero Excel", "✏️ Escribir manualmente"],
                           horizontal=True)

checklist_items = []

if modo_checklist == "📋 Checklist ACCUA predefinida":
    lista_seleccionada = st.selectbox("Selecciona la checklist ACCUA:", list(CHECKLISTS_ACCUA.keys()))
    checklist_items = CHECKLISTS_ACCUA[lista_seleccionada]
    with st.expander(f"👁️ Ver checklist seleccionada ({len(checklist_items)} criterios)"):
        for item in checklist_items:
            st.markdown(f"- {item}")

elif modo_checklist == "📁 Cargar fichero Excel":
    uploaded_file = st.file_uploader("Cargar Checklist (Excel)", type=['xlsx', 'xls'])
    if uploaded_file:
        try:
            if uploaded_file.size > 0:
                import openpyxl as _openpyxl
                import io as _io
                wb_upload = _openpyxl.load_workbook(_io.BytesIO(uploaded_file.read()), data_only=True)
                ws_upload = wb_upload.active

                items_from_excel = []
                skipped_headers = []
                
                for row in ws_upload.iter_rows(min_row=2):
                    cell = row[0]
                    val = str(cell.value).strip() if cell.value else ""
                    if not val or val.lower() in ['información', 'informacion', 'nan', 'check list', 'titulación', 'url', 'check']:
                        continue
                    is_bold = cell.font.bold if cell.font else False
                    if is_bold:
                        skipped_headers.append(val)
                    else:
                        items_from_excel.append(val)
                
                checklist_items.extend(items_from_excel)
                
                if items_from_excel:
                    st.success(f"✅ Se cargaron **{len(items_from_excel)} items** desde el Excel. ({len(skipped_headers)} apartados/secciones ignorados correctamente)")
                    if skipped_headers:
                        with st.expander(f"ℹ️ Secciones detectadas y omitidas ({len(skipped_headers)})"):
                            for h in skipped_headers:
                                st.markdown(f"- _{h[:80]}_")
                else:
                    st.warning("⚠️ No se encontraron items en el archivo Excel. ¿Es el formato esperado?")
            else:
                st.warning("⚠️ El archivo subido está vacío.")
        except Exception as e:
            st.error(f"Error al leer el archivo Excel: {e}")

elif modo_checklist == "✏️ Escribir manualmente":
    checklist_text = st.text_area("Escribe cada requisito en una línea separada:", height=150)
    if checklist_text:
        checklist_items = [line.strip() for line in checklist_text.split('\n') if line.strip()]

# Eliminar duplicados
seen = set()
checklist_items = [x for x in checklist_items if not (x in seen or seen.add(x))]

# --- Botón de Análisis ---
st.markdown("---")
if not api_key:
    st.warning("⚠️ Por favor, introduce tu **Gemini API Key** en el menú lateral para continuar.")
    
analyze_btn = st.button("🚀 Analizar Sitio Web", type="primary",
                        disabled=(not url_input or not checklist_items or not api_key))

if analyze_btn:
    # Guardar URL en historial de sesión
    if "url_history" not in st.session_state:
        st.session_state.url_history = []
    if url_input not in st.session_state.url_history:
        st.session_state.url_history.insert(0, url_input)
        st.session_state.url_history = st.session_state.url_history[:5]

    with st.spinner(f"🕷️ Rastreando sitio web (hasta {max_pages} páginas)..."):
        site_text = crawl_website(url_input, max_pages)
        
    if isinstance(site_text, str) and site_text.startswith("Error"):
        st.error(site_text)
    else:
        st.success("✅ Rastreo web completado con éxito.")
        
        with st.expander("📄 Ver texto extraído (debug)"):
            st.text(site_text[:2000] + "..." if len(site_text) > 2000 else site_text)

        with st.spinner(f"🤖 Analizando cumplimiento con {selected_model}..."):
            audit_results = analyze_compliance(site_text, checklist_items, api_key, selected_model)

        if isinstance(audit_results, dict) and "error" in audit_results:
             st.error(f"Error en el análisis de IA: {audit_results['error']}")
        else:
            st.subheader("📊 Resultados de la Auditoría")
            
            try:
                df_results = pd.DataFrame(audit_results)
                
                if not df_results.empty and 'status' in df_results.columns:
                    total_items = len(df_results)
                    cumple_count = len(df_results[df_results['status'] == 'CUMPLE'])
                    parcial_count = len(df_results[df_results['status'] == 'PARCIAL'])
                    no_cumple_count = len(df_results[df_results['status'] == 'NO CUMPLE'])
                    compliance_rate = (cumple_count / total_items) * 100 if total_items > 0 else 0
                    
                    m1, m2, m3, m4, m5 = st.columns(5)
                    m1.metric("Total criterios", total_items)
                    m2.metric("✅ Cumple", cumple_count)
                    m3.metric("⚠️ Parcial", parcial_count)
                    m4.metric("❌ No cumple", no_cumple_count)
                    m5.metric("% Cumplimiento", f"{compliance_rate:.1f}%")
                    
                    # Resumen visual agrupado
                    st.markdown("---")
                    st.progress(int(compliance_rate) / 100, text=f"Tasa de cumplimiento: {compliance_rate:.1f}%")

                    def color_status(val):
                        if not isinstance(val, str): return ''
                        if 'CUMPLE' in val and 'NO' not in val: return 'color: #a90a2e; font-weight: bold'
                        if 'NO CUMPLE' in val: return 'color: darkred; font-weight: bold'
                        return 'color: #666; font-weight: bold'

                    st.dataframe(
                        df_results.style.map(color_status, subset=['status']),
                        use_container_width=True,
                        column_config={
                            "item": st.column_config.TextColumn("Criterio / Requisito", width="large"),
                            "status": st.column_config.TextColumn("Estado", width="medium"),
                            "evidencia": st.column_config.TextColumn("Evidencia / Justificación", width="large"),
                            "confianza": st.column_config.TextColumn("Confianza", width="small")
                        }
                    )
                    
                    # --- Descargas ---
                    st.markdown("### 📥 Descargar Informe")
                    dcol1, dcol2 = st.columns(2)
                    
                    # Excel
                    with dcol1:
                        try:
                            excel_data = generate_styled_excel(df_results, titulo=titulo_nombre, url=url_input)
                            st.download_button(
                                label="📊 Descargar Excel (.xlsx)",
                                data=excel_data,
                                file_name=f"auditoria_ACCUA_{titulo_nombre.replace(' ','_') if titulo_nombre else 'informe'}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                        except Exception as e:
                           st.error(f"Error al generar el Excel: {e}")

                    # PDF
                    with dcol2:
                        try:
                            pdf_data = generate_pdf_report(df_results, titulo=titulo_nombre, url=url_input, compliance_rate=compliance_rate)
                            if pdf_data:
                                st.download_button(
                                    label="📄 Descargar PDF (Informe completo)",
                                    data=pdf_data,
                                    file_name=f"auditoria_ACCUA_{titulo_nombre.replace(' ','_') if titulo_nombre else 'informe'}.pdf",
                                    mime="application/pdf",
                                    use_container_width=True,
                                )
                            else:
                                st.warning("No se pudo generar el PDF.")
                        except Exception as e:
                            st.error(f"Error al generar el PDF: {e}")
                else:
                    st.warning("⚠️ No se obtuvieron resultados válidos del análisis.")
                    st.write("Datos crudos:", audit_results)
                
            except Exception as e:
                st.error(f"Error al procesar los resultados: {e}")
                st.write("Respuesta cruda:", audit_results)

# Historial de URLs
if "url_history" in st.session_state and st.session_state.url_history:
    with st.sidebar:
        st.markdown("---")
        st.markdown("### 🕐 URLs recientes")
        for url in st.session_state.url_history:
            st.markdown(f"- [{url[:40]}...]({url})" if len(url) > 40 else f"- [{url}]({url})")
